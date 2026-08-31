"""Workbooks Cassandra was not designed around, for honest validation.

The demo model was authored to contain exactly the defects the detectors look
for, which makes it circular as evidence. These three exist to answer questions
the demo model cannot:

  clean_amortisation   A correct workbook. Does Cassandra stay quiet? This is
                       the most important of the three: a tool that invents
                       findings on a healthy file is worse than no tool.

  dept_budget          A different shape entirely, with variance columns and
                       subtotals rather than a projection, carrying two defects
                       written to be plausible rather than convenient.

  saas_projection_v12  The demo model with its defects repaired and one new
                       defect introduced, so the regression sentinel has
                       something real to catch.
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEAD = PatternFill("solid", fgColor="1F3864")
HEADF = Font(color="FFFFFF", bold=True, size=11)
LBL = Font(bold=True, color="1F3864")
TITLE = Font(bold=True, size=14, color="1F3864")
MONEY = "#,##0"
PCT = "0.0%"


def _head(ws, row: int, labels: list[str], col0: int = 2) -> None:
    for i, text in enumerate(labels):
        c = ws.cell(row=row, column=col0 + i, value=text)
        c.fill, c.font = HEAD, HEADF
        c.alignment = Alignment(horizontal="center")


def build_clean_amortisation(path: str) -> str:
    """A correct loan schedule. Nothing here should be reported.

    Deliberately unlike the demo model: one long table, a repeated row formula
    over 24 periods, and functions the demo never uses.
    """
    book = Workbook()
    ws = book.active
    ws.title = "Loan"
    ws["A1"] = "Equipment Loan Amortisation"
    ws["A1"].font = TITLE
    ws.column_dimensions["A"].width = 22
    for col in "BCDEF":
        ws.column_dimensions[col].width = 15

    ws["A3"] = "Principal"
    ws["A3"].font = LBL
    ws["B3"] = 250000
    ws["B3"].number_format = MONEY
    ws["A4"] = "Annual rate"
    ws["A4"].font = LBL
    ws["B4"] = 0.068
    ws["B4"].number_format = PCT
    ws["A5"] = "Term in months"
    ws["A5"].font = LBL
    ws["B5"] = 24
    ws["A6"] = "Monthly payment"
    ws["A6"].font = LBL
    ws["B6"] = "=-PMT($B$4/12,$B$5,$B$3)"
    ws["B6"].number_format = MONEY

    ws["A8"] = "Schedule"
    ws["A8"].font = LBL
    _head(ws, 8, ["Opening", "Payment", "Interest", "Principal", "Closing"], col0=2)

    for i in range(24):
        r = 9 + i
        ws.cell(row=r, column=1, value=f"Month {i + 1}")
        # Every row is the same intent, so the region is homogeneous by design.
        ws.cell(row=r, column=2, value="=E8" if i == 0 else f"=F{r - 1}")
        ws.cell(row=r, column=3, value="=$B$6")
        ws.cell(row=r, column=4, value=f"=B{r}*$B$4/12")
        ws.cell(row=r, column=5, value=f"=C{r}-D{r}")
        ws.cell(row=r, column=6, value=f"=B{r}-E{r}")
        for col in range(2, 7):
            ws.cell(row=r, column=col).number_format = MONEY
    ws["B9"] = "=$B$3"

    ws["A34"] = "Total interest"
    ws["A34"].font = LBL
    ws["D34"] = "=SUM(D9:D32)"
    ws["D34"].number_format = MONEY
    ws["A35"] = "Total repaid"
    ws["A35"].font = LBL
    ws["C35"] = "=SUM(C9:C32)"
    ws["C35"].number_format = MONEY

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    book.save(path)
    return path


def build_dept_budget(path: str) -> str:
    """A budget versus actual sheet, with two defects of a different character.

      Budget!D9  variance formula reversed against the rest of its column
      Budget!C15 subtotal omits the last category row
    """
    book = Workbook()
    ws = book.active
    ws.title = "Budget"
    ws["A1"] = "Marketing Department, FY2027"
    ws["A1"].font = TITLE
    ws.column_dimensions["A"].width = 26
    for col in "BCDE":
        ws.column_dimensions[col].width = 15

    ws["A3"] = "Line"
    ws["A3"].font = LBL
    _head(ws, 3, ["Budget", "Actual", "Variance", "Variance %"])

    rows = [
        ("Paid search", 240000, 268400),
        ("Content and SEO", 120000, 111200),
        ("Events", 185000, 203900),
        ("Brand and creative", 96000, 88700),
        ("Tooling and software", 74000, 79300),
        ("Contract design", 52000, 47600),
    ]
    for i, (name, budget, actual) in enumerate(rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=budget).number_format = MONEY
        ws.cell(row=r, column=3, value=actual).number_format = MONEY
        # DEFECT: row 9 subtracts the other way round, so an overspend on
        # contract design reads as an underspend.
        formula = f"=B{r}-C{r}" if r != 9 else f"=C{r}-B{r}"
        ws.cell(row=r, column=4, value=formula).number_format = MONEY
        ws.cell(row=r, column=5, value=f"=D{r}/B{r}").number_format = PCT

    ws["A12"] = "Totals"
    ws["A12"].font = LBL
    ws["B12"] = "=SUM(B4:B9)"
    ws["C12"] = "=SUM(C4:C9)"
    ws["D12"] = "=SUM(D4:D9)"
    for ref in ("B12", "C12", "D12"):
        ws[ref].number_format = MONEY

    ws["A14"] = "Committed spend"
    ws["A14"].font = LBL
    ws["A15"] = "Total committed"
    # DEFECT: stops at C8 and drops the final category.
    ws["C15"] = "=SUM(C4:C8)"
    ws["C15"].number_format = MONEY

    ws["A17"] = "Budget utilisation"
    ws["A17"].font = LBL
    ws["C17"] = "=C12/B12"
    ws["C17"].number_format = PCT

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    book.save(path)
    return path


def build_v12(path: str) -> str:
    """The demo model with v11 repaired, and one new defect introduced.

    This is what the regression sentinel exists for: everything that was wrong
    last week is fixed, and a single fresh mistake has been made since.
    """
    from .build_workbook import _build_assumptions, _build_pl, _build_revenue

    book = Workbook()
    _build_assumptions(book.active)
    _build_revenue(book.create_sheet("Revenue"))
    _build_pl(book.create_sheet("PL"))

    rev, pl = book["Revenue"], book["PL"]

    # Everything Cassandra found in v11, repaired.
    rev["F8"] = "=SUM(B6:E6)"
    rev["F16"] = "=SUM(B15:E15)"
    rev["D12"] = "=D6*Assumptions!$B$5"
    pl["C8"] = "=C6-C7"
    pl["C13"] = "=Revenue!C12*4"

    # NEW in this revision: somebody typed a one off uplift into Q3 bookings
    # rather than adjusting the assumption, so that quarter no longer computes
    # the way the three beside it do.
    #
    # An earlier draft put the new defect at PL!C5 instead, hardcoding the gross
    # margin. Cassandra did not find it, and the reason is worth keeping: the
    # hardcode detector reports a cell that breaks its region's norm, and the
    # PL column has five different formulas in five rows, so there is no norm to
    # break. Hardcodes are caught where a pattern exists to violate, which is
    # most of a real model and not all of it.
    rev["D6"] = "=D5*Assumptions!$B$6*1.12"

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    book.save(path)
    return path


if __name__ == "__main__":
    for fn, name in (
        (build_clean_amortisation, "clean_amortisation.xlsx"),
        (build_dept_budget, "dept_budget.xlsx"),
        (build_v12, "saas_projection_v12.xlsx"),
    ):
        print("wrote", fn(os.path.join("demo", name)))
