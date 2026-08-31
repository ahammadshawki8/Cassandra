"""Generate the demo workbook: a SaaS three year projection with planted defects.

This is a realistic artifact, not a toy. It has an assumptions sheet feeding a
revenue build feeding a P&L, which is how these models are actually structured,
and the headline ARR figure at the end is the number a founder would put in
front of a board.

Six defects are planted, each drawn from the Panko and Halverson taxonomy and
each representing a class Cassandra hunts. One of them is a deliberate trap:
the obvious repair creates a circular reference, so the Verifier must reject
the first attempt and force a revision. That rejection is the proof that the
loop is real rather than decorative.

Defects planted, by cell:

  Revenue!F8    range omits the final quarter          (Reinhart and Rogoff class)
  Revenue!D12   growth rate hardcoded, breaks the row  (London Whale class)
  PL!C7         opex added rather than subtracted      (sign inversion)
  PL!C10        labelled Net Margin, computes gross    (semantic mismatch)
  PL!C13        references a deleted range             (reference integrity)
  Revenue!F16   trap: naive repair becomes circular    (verifier rejection)
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
LABEL_FONT = Font(bold=True, color="1F3864")
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
MONEY = '#,##0'
PCT = '0.0%'
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

QUARTERS = ["Q1", "Q2", "Q3", "Q4"]


def _header(ws, row: int, labels: list[str], start_col: int = 2) -> None:
    for i, text in enumerate(labels):
        cell = ws.cell(row=row, column=start_col + i, value=text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BOX


def _build_assumptions(ws) -> None:
    ws.title = "Assumptions"
    ws["A1"] = "Operating Assumptions"
    ws["A1"].font = TITLE_FONT
    ws.column_dimensions["A"].width = 34
    for col in "BCDE":
        ws.column_dimensions[col].width = 14

    ws["A3"] = "Driver"
    ws["A3"].font = LABEL_FONT
    _header(ws, 3, ["Value", "Unit"])

    rows = [
        ("Starting customers", 420, "count"),
        ("Quarterly logo growth", 0.18, "rate"),
        ("Average revenue per account", 1450, "USD"),
        ("Quarterly gross churn", 0.055, "rate"),
        ("Gross margin", 0.78, "rate"),
        ("Sales and marketing per quarter", 385000, "USD"),
        ("Research and development per quarter", 512000, "USD"),
        ("General and administrative per quarter", 194000, "USD"),
    ]
    for i, (label, value, unit) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(bold=False)
        cell = ws.cell(row=i, column=2, value=value)
        cell.number_format = PCT if unit == "rate" else MONEY
        cell.border = BOX
        ws.cell(row=i, column=3, value=unit).alignment = Alignment(horizontal="center")


def _build_revenue(ws) -> None:
    ws.title = "Revenue"
    ws["A1"] = "Revenue Build"
    ws["A1"].font = TITLE_FONT
    ws.column_dimensions["A"].width = 34
    for col in "BCDEF":
        ws.column_dimensions[col].width = 15

    ws["A3"] = "FY2027"
    ws["A3"].font = LABEL_FONT
    _header(ws, 3, QUARTERS + ["Total"])

    # Customers. Q1 seeds from assumptions, later quarters compound growth net
    # of churn. Each formula in row 5 shares one intent, so they share one R1C1
    # signature, which is what makes the row a region.
    ws["A5"] = "Customers"
    ws["A5"].font = LABEL_FONT
    ws["B5"] = "=Assumptions!B4"
    for col in "CDE":
        prev = chr(ord(col) - 1)
        ws[f"{col}5"] = f"={prev}5*(1+Assumptions!$B$5)*(1-Assumptions!$B$7)"
    for col in "BCDE":
        ws[f"{col}5"].number_format = "#,##0"

    # Bookings per quarter.
    ws["A6"] = "Bookings"
    ws["A6"].font = LABEL_FONT
    for col in "BCDE":
        ws[f"{col}6"] = f"={col}5*Assumptions!$B$6"
        ws[f"{col}6"].number_format = MONEY

    # DEFECT 1, range omission. The total stops at E and silently drops Q4.
    ws["A8"] = "Total Bookings"
    ws["A8"].font = LABEL_FONT
    ws["F8"] = "=SUM(B6:D6)"
    ws["F8"].number_format = MONEY
    ws["F8"].border = BOX

    # Expansion revenue. Row 12 should reference the assumption like its
    # neighbours do.
    ws["A11"] = "Expansion"
    ws["A11"].font = LABEL_FONT
    ws["A12"] = "Expansion revenue"
    for col in "BCE":
        ws[f"{col}12"] = f"={col}6*Assumptions!$B$5"
        ws[f"{col}12"].number_format = MONEY
    # DEFECT 2, hardcoded constant. Someone typed the growth rate inline.
    ws["D12"] = "=D6*0.18"
    ws["D12"].number_format = MONEY

    # DEFECT 6, the trap. Total Revenue sums the quarterly rows but omits the
    # expansion row. The obvious repair is to widen the range down through row
    # 16, which swallows the total cell itself and creates a circular
    # reference. The Verifier must reject that and force a real fix.
    ws["A15"] = "Quarterly revenue"
    ws["A15"].font = LABEL_FONT
    for col in "BCDE":
        ws[f"{col}15"] = f"={col}6+{col}12"
        ws[f"{col}15"].number_format = MONEY

    ws["A16"] = "Total Revenue"
    ws["A16"].font = LABEL_FONT
    ws["F16"] = "=SUM(B15:D15)"
    ws["F16"].number_format = MONEY
    ws["F16"].border = BOX


def _build_pl(ws) -> None:
    ws.title = "PL"
    ws["A1"] = "Profit and Loss, FY2027"
    ws["A1"].font = TITLE_FONT
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 18

    ws["A3"] = "Line item"
    ws["A3"].font = LABEL_FONT
    _header(ws, 3, ["Note", "FY2027"], start_col=2)

    ws["A4"] = "Revenue"
    ws["C4"] = "=Revenue!F16"
    ws["A5"] = "Cost of revenue"
    ws["C5"] = "=-C4*(1-Assumptions!B8)"
    ws["A6"] = "Gross Profit"
    ws["C6"] = "=C4+C5"

    ws["A7"] = "Total Operating Expenses"
    # DEFECT 3, sign inversion. Opex is added to gross profit rather than
    # subtracted from it, so the model reports profit where there is a loss.
    ws["C7"] = "=Assumptions!B9*4+Assumptions!B10*4+Assumptions!B11*4"

    ws["A8"] = "Operating Income"
    ws["C8"] = "=C6+C7"

    ws["A10"] = "Net Margin"
    # DEFECT 4, semantic mismatch. Labelled net margin, computes gross margin.
    ws["C10"] = "=C6/C4"
    ws["C10"].number_format = PCT

    ws["A12"] = "Headline"
    ws["A12"].font = LABEL_FONT
    ws["A13"] = "Ending ARR"
    # DEFECT 5, reference integrity. Points at a range that no longer exists.
    ws["C13"] = "=Revenue!#REF!*4"

    ws["A15"] = "ARR (board figure)"
    ws["A15"].font = LABEL_FONT
    ws["C15"] = "=Revenue!E6*4"

    for ref in ("C4", "C5", "C6", "C7", "C8", "C15"):
        ws[ref].number_format = MONEY
    for ref in ("C15",):
        ws[ref].font = Font(bold=True, size=12, color="1F3864")
        ws[ref].border = BOX


def build(path: str) -> str:
    """Write the demo workbook and return its path."""
    book = Workbook()
    _build_assumptions(book.active)
    _build_revenue(book.create_sheet("Revenue"))
    _build_pl(book.create_sheet("PL"))
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    book.save(path)
    return path


if __name__ == "__main__":
    target = os.path.join("demo", "saas_projection_v11.xlsx")
    print("wrote", build(target))
