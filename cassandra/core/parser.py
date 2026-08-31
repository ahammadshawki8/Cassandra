"""Workbook parsing: xlsx on disk into the Cassandra cell model.

openpyxl exposes either formulas or cached values, never both in one pass, so
the file is opened twice and the two views are merged.
"""

from __future__ import annotations

import warnings

import openpyxl

from .model import Cell, CellRef, Sheet, Workbook, col_to_letters

warnings.filterwarnings("ignore", module="openpyxl")

# How far to look when hunting for the human label that describes a cell.
_LABEL_SCAN = 8


def parse(path: str) -> Workbook:
    """Parse a workbook into the cell model, merging formulas and cached values."""
    wb_f = openpyxl.load_workbook(path, data_only=False)
    wb_v = openpyxl.load_workbook(path, data_only=True)

    book = Workbook(path=path)
    for name in wb_f.sheetnames:
        ws_f, ws_v = wb_f[name], wb_v[name]
        sheet = Sheet(name=name, max_row=ws_f.max_row or 0, max_col=ws_f.max_column or 0)

        for row in ws_f.iter_rows():
            for c in row:
                raw = c.value
                if raw is None:
                    continue
                is_formula = isinstance(raw, str) and raw.startswith("=")
                ref = CellRef(sheet=name, row=c.row, col=c.column)
                cached = ws_v.cell(row=c.row, column=c.column).value
                sheet.cells[ref.key] = Cell(
                    ref=ref,
                    formula=raw if is_formula else None,
                    value=cached if is_formula else raw,
                    number_format=c.number_format,
                )
        book.sheets[name] = sheet

    _attach_labels(book)
    return book


def _is_text(value: object) -> bool:
    return isinstance(value, str) and not value.startswith("=") and bool(value.strip())


def _attach_labels(book: Workbook) -> None:
    """Find the human readable label describing each formula cell.

    Financial models label rows on the left and columns on top. The Semantic
    Auditor needs this to check that a cell labelled "Net Margin" is not in fact
    computing gross margin.
    """
    for sheet in book.sheets.values():
        for cell in sheet.cells.values():
            if not cell.is_formula:
                continue
            row_label = col_label = None

            for dc in range(1, _LABEL_SCAN + 1):
                col = cell.ref.col - dc
                if col < 1:
                    break
                left = sheet.get(cell.ref.row, col)
                if left and _is_text(left.value):
                    row_label = str(left.value).strip()
                    break

            for dr in range(1, _LABEL_SCAN + 1):
                row = cell.ref.row - dr
                if row < 1:
                    break
                above = sheet.get(row, cell.ref.col)
                if above and _is_text(above.value):
                    col_label = str(above.value).strip()
                    break

            parts = [p for p in (row_label, col_label) if p]
            cell.label = " / ".join(parts) if parts else None
