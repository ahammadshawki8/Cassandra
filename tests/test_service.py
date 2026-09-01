"""Tests for the ways a workbook gets in, and the corrected file that comes out.

These exercise the real code paths. The only thing stubbed is the network call
to Google, because a live test would depend on a third party sheet staying
shared, which is not a property a test suite should rely on. Everything either
side of that call is the code that runs in production.
"""

from __future__ import annotations

import openpyxl
import pytest

from cassandra.core import oracle, parser
from cassandra.service import sources


@pytest.fixture(scope="module")
def workbook_bytes(tmp_path_factory) -> bytes:
    path = tmp_path_factory.mktemp("src") / "model.xlsx"
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Model"
    for i, v in enumerate([10, 20, 30, 40], start=2):
        sheet.cell(row=i, column=1, value=v)
    sheet["A6"] = "Total"
    sheet["B6"] = "=SUM(A2:A4)"
    book.save(path)
    return path.read_bytes()


class TestRejections:
    """Every refusal has to tell the person what to do instead."""

    def test_csv_is_refused_with_the_reason(self):
        with pytest.raises(sources.SourceError) as e:
            sources.from_upload("data.csv", b"a,b\n1,2\n")
        assert "no formulas" in str(e.value)
        assert ".xlsx" in str(e.value)

    def test_legacy_formats_name_the_fix(self):
        with pytest.raises(sources.SourceError) as e:
            sources.from_upload("old.xls", b"whatever")
        assert ".xlsx" in str(e.value)

    def test_unknown_extension(self):
        with pytest.raises(sources.SourceError):
            sources.from_upload("notes.txt", b"hello")

    def test_a_renamed_file_is_caught_by_its_bytes(self):
        """An .xlsx is a zip. A file renamed rather than saved is not one."""
        with pytest.raises(sources.SourceError) as e:
            sources.from_upload("fake.xlsx", b"this is plain text")
        assert "renamed" in str(e.value)

    def test_empty_upload(self):
        with pytest.raises(sources.SourceError):
            sources.from_upload("empty.xlsx", b"")

    def test_oversized_upload(self):
        with pytest.raises(sources.SourceError) as e:
            sources.from_upload("big.xlsx", b"PK" + b"0" * (sources.MAX_BYTES + 1))
        assert "larger than" in str(e.value)


class TestUpload:
    def test_a_real_workbook_lands_on_disk_and_parses(self, workbook_bytes):
        src = sources.from_upload("quarterly model.xlsx", workbook_bytes)
        assert src.origin == "upload"
        book = parser.parse(src.path)
        assert book.get("Model!B6").formula == "=SUM(A2:A4)"

    def test_path_separators_cannot_escape_the_work_directory(self, workbook_bytes):
        src = sources.from_upload("../../etc/passwd.xlsx", workbook_bytes)
        assert ".." not in src.name and "/" not in src.name


class TestGoogleSheets:
    """The import path, with only the call to Google replaced."""

    URL = "https://docs.google.com/spreadsheets/d/1AbCdEfGhIjKlMnOpQrStUvWxYz0123456789/edit#gid=0"

    def test_a_link_that_is_not_sheets_is_rejected(self):
        with pytest.raises(sources.SourceError) as e:
            sources.from_google_sheet("https://example.com/thing")
        assert "not a Google Sheets link" in str(e.value)

    def test_the_export_url_is_built_from_the_sheet_id(self, monkeypatch, workbook_bytes):
        seen = {}

        class Reply:
            status_code = 200
            content = workbook_bytes

        def fake_get(url, **kwargs):
            seen["url"] = url
            return Reply()

        monkeypatch.setattr(sources.requests, "get", fake_get)
        src = sources.from_google_sheet(self.URL)

        assert "1AbCdEfGhIjKlMnOpQrStUvWxYz0123456789" in seen["url"]
        assert seen["url"].endswith("/export?format=xlsx")
        assert src.origin == "google_sheets"

    def test_an_imported_sheet_is_a_workbook_the_auditor_can_read(
        self, monkeypatch, workbook_bytes
    ):
        """The whole point: what comes back is auditable, not just downloaded."""

        class Reply:
            status_code = 200
            content = workbook_bytes

        monkeypatch.setattr(sources.requests, "get", lambda url, **kw: Reply())
        src = sources.from_google_sheet(self.URL)

        book = parser.parse(src.path)
        assert book.get("Model!B6").formula == "=SUM(A2:A4)"
        assert oracle.calculate(src.path, list(book.sheets)).get("Model!B6") == 60.0

    def test_a_private_sheet_explains_how_to_share_it(self, monkeypatch):
        class Reply:
            status_code = 200
            content = b"<html><head><title>Sign in</title></head>"

        monkeypatch.setattr(sources.requests, "get", lambda url, **kw: Reply())
        with pytest.raises(sources.SourceError) as e:
            sources.from_google_sheet(self.URL)
        assert "Anyone with the link" in str(e.value)

    def test_a_forbidden_response_explains_how_to_share_it(self, monkeypatch):
        class Reply:
            status_code = 403
            content = b""

        monkeypatch.setattr(sources.requests, "get", lambda url, **kw: Reply())
        with pytest.raises(sources.SourceError) as e:
            sources.from_google_sheet(self.URL)
        assert "private" in str(e.value)

    def test_a_network_failure_is_reported_not_swallowed(self, monkeypatch):
        def boom(url, **kw):
            raise sources.requests.RequestException("connection reset")

        monkeypatch.setattr(sources.requests, "get", boom)
        with pytest.raises(sources.SourceError) as e:
            sources.from_google_sheet(self.URL)
        assert "connection reset" in str(e.value)


class TestCorrectedCopy:
    """The artifact that makes this a tool rather than a report."""

    def test_corrections_are_written_and_the_original_is_untouched(self, workbook_bytes):
        src = sources.from_upload("model.xlsx", workbook_bytes)
        before = oracle.calculate(src.path, ["Model"]).get("Model!B6")
        assert before == 60.0

        out = sources.corrected_copy(src.path, {"Model!B6": "=SUM(A2:A5)"})
        assert out != src.path

        fixed = parser.parse(out)
        assert fixed.get("Model!B6").formula == "=SUM(A2:A5)"
        assert oracle.calculate(out, ["Model"]).get("Model!B6") == 100.0

        # The file the user uploaded must be exactly as they left it.
        assert oracle.calculate(src.path, ["Model"]).get("Model!B6") == 60.0

    def test_a_missing_original_says_so_rather_than_failing_obscurely(self):
        with pytest.raises(sources.SourceError) as e:
            sources.corrected_copy("no/such/file.xlsx", {"Model!B6": "=1"})
        assert "Run the audit again" in str(e.value)


class TestPathLeaf:
    """Run titles come from paths recorded on whichever machine ran the audit."""

    def test_a_windows_path_is_reduced_on_a_linux_container(self):
        from cassandra.service.app import _leaf

        # os.path.basename only splits on the host separator, so this exact case
        # once put an entire temporary path in the history rail.
        assert _leaf(r"C:\Users\a\AppData\Local\Temp\cassandra\model.xlsx") == "model.xlsx"

    def test_posix_and_bare_names(self):
        from cassandra.service.app import _leaf

        assert _leaf("/tmp/cassandra/model.xlsx") == "model.xlsx"
        assert _leaf("demo/model.xlsx") == "model.xlsx"
        assert _leaf("model.xlsx") == "model.xlsx"

    def test_missing_path_is_empty_not_an_error(self):
        from cassandra.service.app import _leaf

        assert _leaf(None) == ""


class TestLineage:
    """A revision has to be recognised as the same model, or there is no sentinel."""

    def test_version_suffixes_pair_up(self):
        from cassandra.service.store import lineage_key as k

        assert k("saas_projection_v11.xlsx") == k("saas_projection_v12.xlsx")
        assert k("demo/plan_v3.xlsx") == k("/tmp/plan_v9.xlsx")

    def test_the_endings_people_actually_use(self):
        from cassandra.service.store import lineage_key as k

        assert k("Board Model FINAL.xlsx") == "board model"
        assert k("budget (3).xlsx") == "budget"
        assert k("forecast copy.xlsx") == "forecast"
        assert k("plan_2027-03-14.xlsx") == "plan"

    def test_unrelated_models_stay_apart(self):
        from cassandra.service.store import lineage_key as k

        assert k("dept_budget.xlsx") != k("saas_projection_v11.xlsx")

    def test_a_windows_path_reduces_like_any_other(self):
        from cassandra.service.store import lineage_key as k

        assert k(r"C:\Users\a\Temp\saas_projection_v11.xlsx") == "saas_projection"


class TestSettledFindings:
    """A finding a human settled must not come back on the next revision."""

    def test_a_dismissal_is_scoped_to_the_model(self):
        from cassandra.service.store import Store

        store = Store.__new__(Store)
        store._memory, store._claims, store._dismissed = {}, set(), {}
        store._db = None
        import threading

        store._lock = threading.Lock()

        store.dismiss("demo/saas_projection_v11.xlsx", "Revenue!B5")
        # The same address in a different model is untouched.
        assert "Revenue!B5" in store.dismissals("saas_projection_v12.xlsx")
        assert "Revenue!B5" not in store.dismissals("dept_budget.xlsx")

    def test_a_dismissal_can_be_undone(self):
        from cassandra.service.store import Store
        import threading

        store = Store.__new__(Store)
        store._memory, store._claims, store._dismissed = {}, set(), {}
        store._db = None
        store._lock = threading.Lock()

        store.dismiss("model_v1.xlsx", "PL!C8")
        assert store.dismissals("model_v2.xlsx") == {"PL!C8"}
        store.undismiss("model_v2.xlsx", "PL!C8")
        assert store.dismissals("model_v1.xlsx") == set()

    def test_settled_cells_never_reach_the_agents(self, tmp_path):
        """The point of settling: no model call is spent on it at all."""
        import openpyxl
        from cassandra.orchestrator import Auditor

        path = tmp_path / "m.xlsx"
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Model"
        for i, v in enumerate([10, 20, 30, 40], start=2):
            sheet.cell(row=i, column=1, value=v)
        sheet["A6"] = "Total"
        sheet["B6"] = "=SUM(A2:A4)"
        book.save(path)

        # use_agents False keeps this deterministic; the filter runs before the
        # agents either way.
        run = Auditor(use_agents=False, dismissed={"Model!B6"}).audit(str(path))
        assert "Model!B6" in run.settled
        assert all(r.cell != "Model!B6" for r in run.results)


class TestSourceArchive:
    """A corrected download has to survive the container it was audited on.

    Cloud Run scales to zero, so the local file a run was audited from is gone
    the moment the instance recycles. Every stored run's download depended on
    that file still being there.
    """

    def test_the_archive_key_is_scoped_to_the_run(self):
        key = sources.archive_key("abc123", "/tmp/cassandra/model_v11.xlsx")
        assert key == "runs/abc123/model_v11.xlsx"

    def test_a_windows_source_path_still_yields_a_clean_key(self):
        key = sources.archive_key("abc123", r"C:\Users\a\model.xlsx")
        assert key == "runs/abc123/model.xlsx"

    def test_a_nameless_source_still_produces_a_key(self):
        assert sources.archive_key("abc123", "") == "runs/abc123/workbook.xlsx"

    def test_archiving_without_a_bucket_is_a_no_op_not_a_crash(self, tmp_path):
        src = tmp_path / "model.xlsx"
        src.write_bytes(b"PK\x03\x04")
        sources.archive("", "abc123", str(src))

    def test_archiving_a_missing_file_is_a_no_op_not_a_crash(self):
        sources.archive("some-bucket", "abc123", "/no/such/file.xlsx")

    def test_restoring_without_a_bucket_explains_itself(self):
        with pytest.raises(sources.SourceError) as e:
            sources.restore("", "abc123", "model.xlsx")
        assert "no archive bucket" in str(e.value)


class TestPushEndpointGuards:
    """The bucket that receives workbooks is the one that keeps their archives."""

    def test_the_archive_prefix_is_recognisable(self):
        key = sources.archive_key("abc123", "model.xlsx")
        assert key.startswith(f"{sources.RUN_PREFIX}/")

    def test_an_ordinary_upload_is_not_mistaken_for_an_archive(self):
        assert not "saas_projection_v11.xlsx".startswith(f"{sources.RUN_PREFIX}/")


class TestEventReplay:
    """A dashboard connecting late must see one audit, not a blend of two."""

    def _bus_with_two_runs(self):
        from cassandra.service.app import Bus

        bus = Bus()
        for event in [
            {"kind": "woken", "message": "run A"},
            {"kind": "confirmed", "cell": "A1"},
            {"kind": "stored", "run_id": "aaa"},
            {"kind": "woken", "message": "run B"},
            {"kind": "confirmed", "cell": "B2"},
        ]:
            bus.publish(event)
        return bus

    def _drain(self, q):
        out = []
        while not q.empty():
            out.append(q.get_nowait())
        return out

    def test_replay_starts_at_the_most_recent_run(self):
        replayed = self._drain(self._bus_with_two_runs().subscribe())
        cells = [e.get("cell") for e in replayed if e.get("cell")]
        assert cells == ["B2"], "events from the previous run leaked into the replay"
        assert replayed[0]["kind"] == "woken"

    def test_a_bus_that_never_saw_a_run_still_replays(self):
        from cassandra.service.app import Bus

        bus = Bus()
        bus.publish({"kind": "duplicate", "message": "nothing running"})
        assert len(self._drain(bus.subscribe())) == 1


class TestSampleAuditRoute:
    """The sample button is unauthenticated, so it may only reach demo/."""

    @pytest.fixture(scope="class")
    def client(self):
        from fastapi.testclient import TestClient

        from cassandra.service.app import app

        return TestClient(app)

    @pytest.mark.parametrize(
        "path",
        [
            "../cassandra/service/app.py",
            "/etc/passwd",
            "demo/../README.md",
            r"..\..\Windows\win.ini",
        ],
    )
    def test_paths_outside_demo_are_refused(self, client, path):
        body = client.post("/api/audit", json={"path": path}).json()
        assert "error" in body, f"{path} was accepted"
        assert "no such sample workbook" in body["error"]

    def test_a_workbook_that_does_not_exist_is_refused(self, client):
        body = client.post("/api/audit", json={"path": "demo/nope.xlsx"}).json()
        assert "no such sample workbook" in body["error"]
